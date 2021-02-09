from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import datetime
import valid
from valid import Valid
from database import DataBase
validater = Valid()
database = DataBase(user, password, host, db)

user = 'daisuke'
password = 'bskt3B05'
host = 'localhost'
db = 'kancho_news'

today = datetime.date.today()
day = today.day
month = today.month
year = today.year
yymmdd = str(format(valid.today, '%Y%m%d'))

wb = openpyxl.load_workbook("官庁URL.xlsx")
ws = wb["Sheet1"]

def store_list_for_xl(candd):
    list_for_xl.append(candd)
    return list_for_xl

url_list = []

for row in ws.iter_rows(min_row=2,max_row=3) :
    if row[0].value is None:
        break
    value_list = []
    for c in row:
        value_list.append(c.value)
    url_list.append(value_list)

driver_path = "driver/chromedriver_mac87"
driver = webdriver.Chrome(executable_path=driver_path)
driver.implicitly_wait(5)

list_for_xl = []


database.conn_db()

for url in url_list:
    kancho_name = str(url[0])
    kancho_url = str(url[1])
    css_date = str(url[2])
    css_links = str(url[3])

    driver.get(kancho_url)

    date_elem = driver.find_element(By.CSS_SELECTOR, css_date)
    date_text = str(date_elem.text)

    links = driver.find_elements(By.CSS_SELECTOR, css_links)

    for link in links :
        link_text = str(link.text)
        link_url = str(link.get_attribute("href"))

        # link_textのバグ回避のバリデーション
        link_text = validater.remove_space(link_text)

        # 候補の配列
        candd = [kancho_name, date_text, link_text, link_url]

        # ここでselectして保存の可否を決める
        sql = f'SELECT link_text FROM data_list WHERE link_text="{candd[2]}" LIMIT 1 ;'
        selects = database.select(sql, candd)

        sql = f'INSERT INTO data_list (kancho_name,date_text,link_text,link_url) VALUES ("{candd[0]}","{candd[1]}","{candd[2]}","{candd[3]}");'

        if not selects:
            insert = database.insert(sql, candd)
            list_for_xl = store_list_for_xl(candd)
        else:
            for select in selects:
                if candd[2] in str(select): # 型一致が必要
                    print(f"This news is arleady exists: [{candd[0]}: {candd[2]}]")
                else:
                    insert = database.insert(sql, candd)
                    list_for_xl = store_list_for_xl(candd)

driver.quit()

database.disconn_db()
# print(list_for_xl)
# exit()


wb_new = openpyxl.Workbook()
ws_new = wb_new.worksheets[0]

row_num = 1

for data in list_for_xl :
    ws_new.cell(row_num, 1).value = data[0]
    ws_new.cell(row_num, 2).value = data[1]
    ws_new.cell(row_num, 3).value = data[2]
    ws_new.cell(row_num, 4).value = data[3]

    row_num += 1

wb_new.save("data/" + yymmdd + "官庁新着情報.xlsx")
